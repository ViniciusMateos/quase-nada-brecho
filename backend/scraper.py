"""
Ponte com o worker `brecho-tracker` (Playwright) que raspa o Instagram do brechó.

Responsabilidades:
  - montar a linha de comando do worker a partir dos params da API
  - salvar os cookies do IG (sessão) no dir do worker
  - depois de uma raspagem, importar a planilha gerada (brecho_tracker.xlsx)
    pra dentro do SQLite (peças com origem='scraper'), pro dashboard refletir
    as vendas reais.

O worker em si (main.py, iglib.py, parser.py, planilha.py) é o mesmo que rodava
no Quase Nada Bots — foi movido pra cá sem alteração.
"""
import json
from datetime import date, datetime

import settings

COOKIES_FILE = "imported_cookies.json"
PLANILHA = "brecho_tracker.xlsx"

# posição das colunas na aba "peças" (mesmo layout gerado pelo worker)
COL = {"code": 0, "drop": 1, "item": 3, "nome": 4, "compra": 5, "venda": 6,
       "tamanho": 7, "largura": 8, "comprimento": 9, "circunferencia": 10,
       "condicao": 11, "vendida": 12, "url": 13, "imagem_url": 15, "numero": 16}


def worker_dir():
    return settings.WORKER_DIR


def montar_cmd(params):
    """Traduz params (dict da API) nos argumentos da CLI do worker."""
    p = params or {}
    if p.get("import_cookies"):                 # conectar IG: só importa e sai
        return ["main.py", "--import-cookies", str(p["import_cookies"])]
    args = ["main.py"]
    if p.get("dry_run"):
        args.append("--dry-run")
    if p.get("full"):
        args.append("--full")
    if p.get("rematch"):
        args.append("--rematch")
    return args


def salvar_cookies(cookies):
    """Grava os cookies (lista, formato Cookie-Editor) no dir do worker. Devolve o
    nome do arquivo (relativo — a run roda com cwd = dir do worker)."""
    (settings.WORKER_DIR / COOKIES_FILE).write_text(
        json.dumps(cookies, ensure_ascii=False, indent=2), encoding="utf-8")
    return COOKIES_FILE


# ─────────────────── import da planilha → SQLite ─────────────────
def _data_iso(v):
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    return str(v) if v else None


def _limpo(v):
    return None if v in (None, "", "N/A") else v


def _cel_int(r, i):
    """Lê uma célula como int (None se vazia ou fora do range — planilha antiga sem a coluna)."""
    if i < len(r) and r[i] not in (None, ""):
        try:
            return int(r[i])
        except (TypeError, ValueError):
            return None
    return None


def _post_url(code):
    return f"https://www.instagram.com/p/{code}/"


def ler_planilha():
    """Lê a aba 'peças' do brecho_tracker.xlsx. Devolve lista de dicts (ou [])."""
    import openpyxl
    path = settings.WORKER_DIR / PLANILHA
    if not path.exists():
        return []
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if "peças" not in wb.sheetnames:
        wb.close()
        return []
    ws = wb["peças"]
    linhas = ws.iter_rows(values_only=True)
    next(linhas, None)                          # pula cabeçalho
    out = []
    for r in linhas:
        if not r or not r[COL["code"]]:
            continue
        out.append({
            "code": str(r[COL["code"]]),
            "drop": _data_iso(r[COL["drop"]]),
            "item": r[COL["item"]],
            "nome": r[COL["nome"]],
            "compra": r[COL["compra"]],
            "venda": r[COL["venda"]],
            "tamanho": r[COL["tamanho"]],
            "largura": r[COL["largura"]],
            "comprimento": r[COL["comprimento"]],
            "circunferencia": r[COL["circunferencia"]],
            "condicao": r[COL["condicao"]],
            "vendida": str(r[COL["vendida"]]).strip().lower() == "sim",
            "imagem_url": _limpo(r[COL["imagem_url"]]) or _post_url(r[COL["code"]]),
            "numero": _cel_int(r, COL["numero"]),   # #p<num> pro match no app
        })
    wb.close()
    return out


_VAZIO = {"novas": 0, "atualizadas": 0, "reconciliadas": 0, "recem_vendidas": 0, "eventos": [], "total": 0}


def _ler_pecas_dump():
    """Peças parseadas que o worker entregou (output/pecas.json)."""
    import json
    f = settings.WORKER_DIR / "output" / "pecas.json"
    if not f.exists():
        return []
    try:
        return json.loads(f.read_text(encoding="utf-8"))
    except Exception:
        return []


def importar():
    """Aplica as peças raspadas no app (SQLite) — o app é a fonte da verdade. Devolve stats."""
    import pecas
    items = _ler_pecas_dump()
    return pecas.upsert_scraper(items) if items else dict(_VAZIO)


def preview_import():
    """DRY-RUN: SIMULA o upsert no app com rollback — devolve o que MUDARIA (casadas por #p,
    novas, atualizadas, venderam) sem gravar nada."""
    import pecas
    items = _ler_pecas_dump()
    return pecas.upsert_scraper(items, preview=True) if items else dict(_VAZIO)


def sincronizar_planilha():
    """Reescreve a planilha do worker A PARTIR DO APP (fonte da verdade) — pra ficarem
    SINCRONIZADAS. Peça manual (sem code) entra com code='p<num>'. O worker usa isso só pra
    calcular o boundary (onde parar de descer o feed); quem reconcilia é o app. Chamado antes
    de cada raspagem e depois de importar."""
    import json
    from datetime import datetime
    import openpyxl
    import pecas
    ps = pecas.listar()["pecas"]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "peças"
    ws.append(["code", "drop", "imagem", "item", "nome", "compra", "venda", "tamanho",
               "largura", "comprimento", "circunferencia", "condicao", "vendida", "url",
               "atualizado_em", "imagem_url", "numero", "so_manual"])
    agora = datetime.now().strftime("%Y-%m-%d %H:%M")
    for p in ps:
        real = p.get("code")
        code = real or (f"p{p['num']}" if p.get("num") else None)
        if not code:
            continue
        circ = None
        try:
            for m in json.loads(p.get("medida") or "[]"):
                if "circ" in (m.get("tipo") or "").lower():
                    circ = m.get("valor")
        except Exception:
            pass
        ws.append([
            code, p.get("postado_em"), None, p.get("item"), p.get("nome"),
            p.get("compra"), p.get("venda"), p.get("tamanho"),
            p.get("largura"), p.get("comprimento"), circ, p.get("condicao"),
            "sim" if p.get("vendida") else "não",
            _post_url(real) if real else None,
            agora, p.get("imagem_url"), p.get("num"), 1 if p.get("so_manual") else 0,
        ])
    wb.save(str(settings.WORKER_DIR / PLANILHA))
    return len(ps)
