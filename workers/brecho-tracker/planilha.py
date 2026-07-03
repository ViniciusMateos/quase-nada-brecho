"""
Motor da planilha do brecho-tracker.

A planilha de saída (brecho_tracker.xlsx) É o banco de dados — chaveada pelo `code`
do post (1 post = 1 peça). Duas abas:
  • "Dados Gerais" (1ª, principal): dashboard com KPIs via FÓRMULAS que apontam p/ peças.
  • "peças" (2ª): a tabela detalhada (foto no hover, datas DD/MM, link, zebra).

A cada run:
  • preserva o que é MANUAL (compra)
  • atualiza o que é RASPADO (vendida, medidas, condição, nome, preço, imagem)

Também lê a planilha ANTIGA (quasenadabrecho.xlsx) pra, no 1º run, puxar
preço/custo das peças já vendidas (match best-effort por drop+item+tamanho).
"""
import glob
import json
import os
import re
import shutil
import tempfile
import unicodedata
import zipfile
from collections import defaultdict
from datetime import datetime

import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage, ImageFont

import config
from iglib import log

# fonte p/ medir a largura real do texto (Arial 12 ≈ 16px @ 96dpi) e calcular o autofit
try:
    _FONTE_MEDIDA = ImageFont.truetype("arial.ttf", 16)
except Exception:
    _FONTE_MEDIDA = None

_CACHE_ANTIGA = os.path.join(config.OUTPUT_DIR, "antiga_cache.json")

# Ordem das colunas da aba 'peças' (a 1ª, code, é a chave).
# 'imagem' = célula com a foto no hover; 'imagem_url' fica por último, OCULTA.
COLUNAS = [
    "code", "drop", "imagem", "item", "nome", "compra", "venda", "tamanho",
    "largura", "comprimento", "circunferencia", "condicao", "vendida", "url",
    "atualizado_em", "imagem_url",
]
COL_OCULTAS = ("imagem_url",)

# ── identidade visual ──
LARANJA = "FF8234"
CINZA_ZEBRA = "F2F2F2"
AZUL_LINK = "0563C1"
_thin = Side(style="thin", color="D9D9D9")
BORDA = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
_MOEDA = 'R$ #,##0.00'

# ── fontes: tudo em Arial 12 ──
FONTE = "Arial"
F = Font(name=FONTE, size=12)
F_BOLD = Font(name=FONTE, size=12, bold=True)
F_BOLD_BRANCO = Font(name=FONTE, size=12, bold=True, color="FFFFFF")
F_LINK = Font(name=FONTE, size=12, color=AZUL_LINK, underline="single")


def _norm(s):
    return str(s).strip().lower() if s is not None else ""


# tipos de calçado que no IG aparecem só pela marca (sem a palavra "tênis")
_ITEM_SINON = {
    "short": "shorts", "jort": "jorts", "sueter": "suter", "sueteres": "suter",
    "nike": "tenis", "adidas": "tenis", "mizuno": "tenis", "puma": "tenis",
    "vans": "tenis", "olympikus": "tenis", "fila": "tenis", "asics": "tenis",
    "newbalance": "tenis", "new balance": "tenis", "reebok": "tenis", "oakley": "oculos",
}


def _item_norm(s):
    """Normaliza o item p/ match: minúsculo, SEM acento, com sinônimos (suéter→suter,
    short→shorts, nike/adidas/…→tenis)."""
    s = str(s or "").strip().lower()
    s = "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")
    return _ITEM_SINON.get(s, s)


def _num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _fmt_valor(v):
    """Formata um valor monetário p/ o log (None/'' → 'sem preço')."""
    if v in (None, ""):
        return "sem preço"
    try:
        return f"R$ {float(v):.2f}"
    except (TypeError, ValueError):
        return str(v)


def backfill_antiga(db, antiga):
    """Preenche compra/venda FALTANTES no db usando a planilha antiga.

    Casa drop-a-drop (as datas batem): dentro de cada drop, pontua cada par
    (peça nova × peça antiga) por item+tamanho+comprimento+preço e atribui de forma
    gulosa (cada peça antiga usada 1x). Só preenche o que está vazio — nunca
    sobrescreve preço de legenda nem custo digitado à mão. Retorna nº de valores preenchidos.
    """
    od = defaultdict(list)
    for a in antiga:
        od[a["drop"]].append({"item": _item_norm(a["item"]), "tam": _norm(a["tamanho"]),
                              "comp": _num(a["comprimento"]), "compra": a["compra"],
                              "venda": _num(a["venda"])})
    nd = defaultdict(list)
    for l in db.values():
        nd[l.get("drop")].append(l)

    preenchidos = 0
    for drop, itens in nd.items():
        cands = od.get(drop, [])
        if not cands:
            continue
        pares = []
        for ni, l in enumerate(itens):
            n_item, n_tam = _item_norm(l.get("item")), _norm(l.get("tamanho"))
            n_comp, n_venda = _num(l.get("comprimento")), _num(l.get("venda"))
            for oi, o in enumerate(cands):
                s = 3 if n_item == o["item"] else 0
                if n_tam and n_tam == o["tam"]:
                    s += 2
                if n_comp is not None and o["comp"] is not None and abs(n_comp - o["comp"]) <= 2:
                    s += 2
                if n_venda is not None and o["venda"] is not None and abs(n_venda - o["venda"]) < 1:
                    s += 2
                if s >= 3:
                    pares.append((s, ni, oi))
        pares.sort(reverse=True)
        usou_n, usou_o = set(), set()
        for s, ni, oi in pares:
            if ni in usou_n or oi in usou_o:
                continue
            usou_n.add(ni); usou_o.add(oi)
            l, o = itens[ni], cands[oi]
            if l.get("compra") in (None, "") and o["compra"] not in (None, ""):
                l["compra"] = o["compra"]; preenchidos += 1
            if l.get("venda") in (None, "") and o["venda"] is not None:
                l["venda"] = o["venda"]; preenchidos += 1
    return preenchidos


# ───────────────────── planilha ANTIGA (referência) ─────────────────────
def carregar_planilha_antiga():
    """Lê a quasenadabrecho.xlsx e indexa por (drop, item, tamanho) p/ match.

    A planilha antiga é pesada (~10 MB) e lenta; cacheamos o parse num JSON leve.
    """
    if not os.path.exists(config.PLANILHA):
        return []
    try:
        if (os.path.exists(_CACHE_ANTIGA)
                and os.path.getmtime(_CACHE_ANTIGA) >= os.path.getmtime(config.PLANILHA)):
            with open(_CACHE_ANTIGA, encoding="utf-8") as f:
                return json.load(f)
    except Exception:
        pass

    log.info("Carregando planilha antiga (~10 MB com fotos — pode levar ~1 min, só nesta vez)…")
    wb = openpyxl.load_workbook(config.PLANILHA, read_only=True, data_only=True)
    if config.ABA_PECAS not in wb.sheetnames:
        return []
    ws = wb[config.ABA_PECAS]
    linhas = list(ws.iter_rows(min_row=config.HEADER_LINHA, values_only=True))
    if not linhas:
        return []
    cab = [_norm(c) for c in linhas[0]]

    def col(nome):
        return cab.index(nome) if nome in cab else None
    ci = {k: col(k) for k in ("item", "tamanho", "largura", "comprimento",
                              "venda", "compra", "drop")}
    out = []
    for row in linhas[1:]:
        if ci["item"] is None or row[ci["item"]] in (None, "", "item"):
            continue

        def g(k):
            i = ci[k]
            return row[i] if i is not None and i < len(row) else None
        drop = g("drop")
        if isinstance(drop, datetime):
            drop = drop.strftime("%Y-%m-%d")
        elif drop is not None:
            drop = str(drop)[:10]
        out.append({"drop": drop, "item": _norm(g("item")), "tamanho": _norm(g("tamanho")),
                    "largura": g("largura"), "comprimento": g("comprimento"),
                    "venda": g("venda"), "compra": g("compra")})
    wb.close()
    try:
        os.makedirs(config.OUTPUT_DIR, exist_ok=True)
        with open(_CACHE_ANTIGA, "w", encoding="utf-8") as f:
            json.dump(out, f, ensure_ascii=False)
        log.info("Planilha antiga indexada (%d peças) — cache salvo.", len(out))
    except Exception:
        pass
    return out


def _dias_entre(d1, d2):
    try:
        a = datetime.strptime(str(d1)[:10], "%Y-%m-%d")
        b = datetime.strptime(str(d2)[:10], "%Y-%m-%d")
        return abs((a - b).days)
    except (ValueError, TypeError):
        return None


def achar_na_antiga(peca, antiga, tol_dias=2):
    """Best-effort: linha da antiga que corresponde à peça (item + drop ±tol; depois
    tamanho exato → comprimento ±2cm → único)."""
    drop, item, tam = peca.get("drop"), _norm(peca.get("item")), _norm(peca.get("tamanho"))
    cand = [r for r in antiga if r["item"] == item
            and (_dias_entre(r["drop"], drop) is not None)
            and _dias_entre(r["drop"], drop) <= tol_dias]
    if not cand:
        return None
    if tam:
        exato = [r for r in cand if r["tamanho"] == tam]
        if len(exato) == 1:
            return exato[0]
        if exato:
            cand = exato
    comp = peca.get("comprimento")
    if comp is not None:
        perto = [r for r in cand if r["comprimento"] is not None
                 and abs(float(r["comprimento"]) - float(comp)) <= 2]
        if len(perto) == 1:
            return perto[0]
        if perto:
            cand = perto
    return cand[0] if len(cand) == 1 else None


# ───────────────────── planilha de SAÍDA (banco) ─────────────────────
def carregar_db():
    """Lê a brecho_tracker.xlsx existente → dict {code: linha}. Vazio se não existe."""
    if not os.path.exists(config.PLANILHA_SAIDA):
        return {}
    wb = openpyxl.load_workbook(config.PLANILHA_SAIDA, data_only=True)
    if config.ABA_PECAS not in wb.sheetnames:
        return {}
    ws = wb[config.ABA_PECAS]
    header_row = None
    for r in range(1, min(ws.max_row, 12) + 1):
        if _norm(ws.cell(r, 1).value) == "code":
            header_row = r
            break
    if header_row is None:
        return {}
    cab = [_norm(ws.cell(header_row, c).value) for c in range(1, len(COLUNAS) + 1)]
    db = {}
    for r in range(header_row + 1, ws.max_row + 1):
        code = ws.cell(r, 1).value
        if not code:
            continue
        linha = {}
        for c, nome in enumerate(cab, start=1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.strip().upper() == "N/A":   # N/A é só visual
                v = None
            elif nome == "drop" and isinstance(v, datetime):
                v = v.strftime("%Y-%m-%d")
            elif nome == "atualizado_em" and isinstance(v, datetime):
                v = v.strftime("%Y-%m-%d %H:%M")
            linha[nome] = v
        db[str(code)] = linha
    wb.close()
    return db


def _to_sim_nao(v):
    return "sim" if v else "não"


def reconciliar(db, pecas_raspadas, antiga):
    """Funde as peças raspadas no db (dict por code). Devolve (db, stats)."""
    agora = datetime.now().strftime("%Y-%m-%d %H:%M")
    novas = atualizadas = recem_vendidas = 0

    for p in pecas_raspadas:
        code = p["code"]
        atual = db.get(code)
        if atual is None:                      # ── peça nova ──
            linha = {k: None for k in COLUNAS}
            linha.update({
                "code": code, "drop": p["drop"], "item": p["item"], "nome": p["nome"],
                "tamanho": p["tamanho"], "largura": p["largura"],
                "comprimento": p["comprimento"], "circunferencia": p["circunferencia"],
                "condicao": p["condicao"], "venda": p["preco"],
                "vendida": _to_sim_nao(p["vendida"]), "url": p["url"],
                "imagem_url": p.get("imagem_url"), "atualizado_em": agora,
            })
            # casa com a antiga p/ trazer o que o IG não tem (custo; e preço se vendida)
            if antiga:
                m = achar_na_antiga(p, antiga)
                if m:
                    if m.get("compra") not in (None, ""):
                        linha["compra"] = m.get("compra")
                    if linha["venda"] in (None, "") and m.get("venda") not in (None, ""):
                        linha["venda"] = m.get("venda")
            db[code] = linha
            novas += 1
            log.info("  ADICIONADA: %s — %s",
                     linha.get("nome") or code, _fmt_valor(linha.get("venda")))
            continue

        # ── peça conhecida: atualiza raspado, preserva manual (compra) ──
        era_vendida = _norm(atual.get("vendida")) == "sim"
        venda_antes = atual.get("venda")
        cond_antes = atual.get("condicao")
        atual["nome"] = p["nome"] or atual.get("nome")
        atual["item"] = atual.get("item") or p["item"]
        atual["tamanho"] = p["tamanho"] or atual.get("tamanho")
        for k in ("largura", "comprimento", "circunferencia", "condicao"):
            if p[k] is not None:
                atual[k] = p[k]
        if p["preco"] is not None:             # preço enquanto disponível (retém o último)
            atual["venda"] = p["preco"]
        atual["vendida"] = _to_sim_nao(p["vendida"])
        atual["url"] = p["url"]
        if p.get("imagem_url"):
            atual["imagem_url"] = p["imagem_url"]
        atual["atualizado_em"] = agora

        # detecta e loga o que mudou (preço de quanto→quanto, vendeu, condição)
        mud = []
        if _num(atual.get("venda")) != _num(venda_antes):
            mud.append(f"preço {_fmt_valor(venda_antes)} → {_fmt_valor(atual.get('venda'))}")
        if p["vendida"] and not era_vendida:
            mud.append("VENDEU (estava disponível)")
            recem_vendidas += 1
        elif not p["vendida"] and era_vendida:
            mud.append("voltou a disponível")
        if cond_antes is not None and atual.get("condicao") != cond_antes:
            mud.append(f"condição {cond_antes} → {atual.get('condicao')}")
        if mud:
            log.info("  ATUALIZADA: %s — %s", atual.get("nome") or code, " · ".join(mud))
            atualizadas += 1

    return db, {"novas": novas, "atualizadas": atualizadas, "recem_vendidas": recem_vendidas}


def boundary_disponivel(db):
    """Drop mais antigo que ainda tem peça disponível. None se nenhuma."""
    drops = [l.get("drop") for l in db.values()
             if _norm(l.get("vendida")) != "sim" and l.get("drop")]
    return min(drops) if drops else None


def _kpis(db):
    """KPIs em Python (p/ os logs do run). A planilha usa fórmulas vivas próprias."""
    def num(v):
        try:
            return float(v)
        except (TypeError, ValueError):
            return 0.0
    vendidas = [l for l in db.values() if _norm(l.get("vendida")) == "sim"]
    todas = list(db.values())
    fat = sum(num(l.get("venda")) for l in vendidas)
    proj = sum(num(l.get("venda")) for l in todas)
    gastos_todas = sum(num(l.get("compra")) for l in todas)
    custo_vendidas = sum(num(l.get("compra")) for l in vendidas)
    return {"faturamento total": fat, "projeção faturamento total": proj,
            "gastos totais": gastos_todas, "custo das vendidas": custo_vendidas,
            "lucro líquido": fat - custo_vendidas,
            "peças": len(todas), "vendidas": len(vendidas),
            "disponíveis": len(todas) - len(vendidas)}


# ──────────────────────────── escrita ────────────────────────────
def salvar(db):
    """Escreve a brecho_tracker.xlsx (Dados Gerais + peças + gastos) e injeta o hover."""
    wb = openpyxl.Workbook()
    ws_dados = wb.active
    ws_dados.title = config.ABA_DADOS
    _aba_dados_gerais(ws_dados, db)

    ws_pecas = wb.create_sheet(config.ABA_PECAS)
    imagens = _aba_pecas(ws_pecas, db)

    wb.save(config.PLANILHA_SAIDA)

    if imagens:                                # foto no hover (best-effort)
        try:
            _injetar_hover(config.PLANILHA_SAIDA, imagens)
        except Exception as e:
            log.warning("Hover de imagem não aplicado (%s). A planilha está OK, só sem a "
                        "prévia ao passar o mouse.", e)
    return _kpis(db)


_PCT = "0.0%"


def _aba_dados_gerais(ws, db):
    """Dashboard: Resumo Geral (financeiro + %), Estoque, e quebra POR ANO — tudo via
    FÓRMULAS que apontam p/ a aba 'peças' (atualiza sozinho)."""
    laranja = PatternFill("solid", fgColor=LARANJA)
    cinza = PatternFill("solid", fgColor=CINZA_ZEBRA)
    P = config.ABA_PECAS
    cl = {n: get_column_letter(i + 1) for i, n in enumerate(COLUNAS)}   # letras dinâmicas
    Lv, Lc, Lvd, Lcode, Ld = cl["venda"], cl["compra"], cl["vendida"], cl["code"], cl["drop"]

    ws.merge_cells("A1:G1")
    t = ws.cell(1, 1, "QUASE NADA BRECHÓ — DADOS GERAIS")
    t.font = F_BOLD_BRANCO; t.fill = laranja
    t.alignment = Alignment(horizontal="center", vertical="center")
    for c in range(2, 8):
        ws.cell(1, c).fill = laranja
    ws.row_dimensions[1].height = 26

    def secao(r, titulo, ncols=3):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
        s = ws.cell(r, 1, titulo); s.font = F_BOLD_BRANCO; s.fill = laranja
        s.alignment = Alignment(horizontal="left", vertical="center")
        for c in range(2, ncols + 1):
            ws.cell(r, c).fill = laranja

    def linha(r, lab, formula, fmt, desc):
        a = ws.cell(r, 1, lab); a.font = F_BOLD; a.border = BORDA
        v = ws.cell(r, 2, formula); v.font = F; v.number_format = fmt
        v.border = BORDA; v.alignment = Alignment(horizontal="right")
        d = ws.cell(r, 3, desc); d.font = F; d.border = BORDA
        d.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    secao(3, "RESUMO GERAL")
    linha(4, "Faturamento total", f"=SUMIF('{P}'!{Lvd}:{Lvd},\"sim\",'{P}'!{Lv}:{Lv})", _MOEDA,
          "Preço das peças já vendidas (grana que entrou).")
    linha(5, "Projeção de faturamento", f"=SUM('{P}'!{Lv}:{Lv})", _MOEDA,
          "Preço de TODAS as peças (se vender tudo).")
    linha(6, "Gastos totais", f"=SUM('{P}'!{Lc}:{Lc})", _MOEDA,
          "Custo de TODAS as peças (investimento total).")
    linha(7, "Custo das vendidas", f"=SUMIF('{P}'!{Lvd}:{Lvd},\"sim\",'{P}'!{Lc}:{Lc})", _MOEDA,
          "Custo só das vendidas.")
    linha(8, "Lucro líquido", "=B4-B7", _MOEDA,
          "Faturamento − custo das vendidas (lucro real do que saiu).")
    linha(9, "Retorno do investimento", "=IFERROR(B8/B7,0)", _PCT,
          "Quanto o custo das vendidas rendeu de lucro.")
    linha(10, "Margem de lucro", "=IFERROR(B8/B4,0)", _PCT,
          "Lucro sobre o faturamento.")

    secao(12, "ESTOQUE")
    linha(13, "Total de peças", f"=COUNTA('{P}'!{Lcode}:{Lcode})-1", "0", "Peças no catálogo.")
    linha(14, "Vendidas", f"=COUNTIF('{P}'!{Lvd}:{Lvd},\"sim\")", "0", "Já venderam.")
    linha(15, "Disponíveis", "=B13-B14", "0", "Ainda à venda.")
    linha(16, "Taxa de venda", "=IFERROR(B14/B13,0)", _PCT, "% das peças que já venderam.")

    # ───── POR ANO ─────
    anos = sorted({int(str(l["drop"])[:4]) for l in db.values()
                   if l.get("drop") and str(l["drop"])[:4].isdigit()})
    r0 = 18
    secao(r0, "POR ANO", ncols=7)
    cabec = ["Ano", "Peças", "Vendidas", "Faturamento", "Custo", "Lucro", "Margem"]
    for ci, h in enumerate(cabec, start=1):
        cell = ws.cell(r0 + 1, ci, h); cell.font = F_BOLD_BRANCO; cell.fill = laranja
        cell.alignment = Alignment(horizontal="center"); cell.border = BORDA
    drng = f"'{P}'!{Ld}:{Ld}"
    for i, ano in enumerate(anos):
        r = r0 + 2 + i
        ini, fim = f'">="&DATE({ano},1,1)', f'"<"&DATE({ano + 1},1,1)'
        vals = [
            (ano, "0"),
            (f"=COUNTIFS({drng},{ini},{drng},{fim})", "0"),
            (f"=COUNTIFS('{P}'!{Lvd}:{Lvd},\"sim\",{drng},{ini},{drng},{fim})", "0"),
            (f"=SUMIFS('{P}'!{Lv}:{Lv},'{P}'!{Lvd}:{Lvd},\"sim\",{drng},{ini},{drng},{fim})", _MOEDA),
            (f"=SUMIFS('{P}'!{Lc}:{Lc},'{P}'!{Lvd}:{Lvd},\"sim\",{drng},{ini},{drng},{fim})", _MOEDA),
            (f"=D{r}-E{r}", _MOEDA),
            (f"=IFERROR(F{r}/D{r},0)", _PCT),
        ]
        for ci, (v, fmt) in enumerate(vals, start=1):
            cell = ws.cell(r, ci, v); cell.font = F; cell.number_format = fmt; cell.border = BORDA
            if ci == 1:
                cell.alignment = Alignment(horizontal="center")
            if i % 2 == 1:
                cell.fill = cinza

    for letra, w in {"A": 26, "B": 16, "C": 46, "D": 13, "E": 13, "F": 13, "G": 11}.items():
        ws.column_dimensions[letra].width = w
    ws.sheet_view.showGridLines = False


def _aba_pecas(ws, db):
    """Tabela detalhada. Retorna lista (row, col, caminho_thumb) p/ o hover."""
    laranja = PatternFill("solid", fgColor=LARANJA)
    zebra = PatternFill("solid", fgColor=CINZA_ZEBRA)
    col = {n: i + 1 for i, n in enumerate(COLUNAS)}

    for n, c in col.items():                   # cabeçalho (linha 1)
        cell = ws.cell(1, c, n)
        cell.font = F_BOLD_BRANCO; cell.fill = laranja
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDA

    linhas = sorted(db.values(), key=lambda l: (str(l.get("drop") or ""), str(l.get("code"))))
    imagens = []
    di = 2
    for idx, l in enumerate(linhas):
        r = di + idx
        zebrar = (idx % 2 == 1)
        for n, c in col.items():
            cell = ws.cell(r, c)
            val = l.get(n)
            if n == "drop" and val:
                try:
                    cell.value = datetime.strptime(str(val)[:10], "%Y-%m-%d")
                    cell.number_format = "DD/MM/YYYY"
                except ValueError:
                    cell.value = val
                cell.font = F
            elif n == "atualizado_em" and val:
                try:
                    cell.value = datetime.strptime(str(val)[:16], "%Y-%m-%d %H:%M")
                    cell.number_format = "DD/MM/YYYY HH:MM"
                except ValueError:
                    cell.value = val
                cell.font = F
            elif n == "url" and val:           # hyperlink nativo (abre a URL ao clicar)
                cell.value = "ver post"
                cell.hyperlink = val
                cell.font = F_LINK
            elif n == "imagem":
                cell.value = "📷"
                cell.font = F
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.value = val if val not in (None, "") else "N/A"
                cell.font = F
            if n in ("compra", "venda") and val not in (None, ""):
                cell.number_format = _MOEDA
            cell.border = BORDA
            if zebrar:
                cell.fill = zebra

        # foto no hover: comentário com a miniatura
        code = l.get("code")
        thumb = os.path.join(config.IMAGENS_DIR, f"{code}.jpg") if code else None
        if thumb and os.path.exists(thumb):
            try:
                tw, th = PILImage.open(thumb).size
                lado = 180
                w, h = (lado * tw / th, lado) if th >= tw else (lado, lado * th / tw)
                cm = Comment(" ", "brecho-tracker", height=int(h), width=int(w))
                ws.cell(r, col["imagem"]).comment = cm
                imagens.append((r, col["imagem"], thumb))
            except Exception:
                pass

    ws.freeze_panes = ws.cell(2, 1)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUNAS))}{max(1, len(linhas) + 1)}"
    _ajustar_larguras(ws, col, linhas)
    return imagens


def _texto_exibido(n, val):
    """O que de fato aparece na célula (p/ medir a largura — datas/moeda/link têm
    formato próprio, diferente do valor cru)."""
    if n == "drop":
        return "00/00/0000" if val not in (None, "") else ""
    if n == "atualizado_em":
        return "00/00/0000 00:00" if val not in (None, "") else ""
    if n == "url":
        return "ver post" if val not in (None, "") else ""
    if n == "imagem":
        return "x"
    if val in (None, ""):
        return "N/A"
    if n in ("compra", "venda"):
        try:
            return f"R$ {float(val):,.2f}"
        except (TypeError, ValueError):
            return str(val)
    return str(val)


def _px(texto):
    """Largura do texto em pixels na fonte Arial 12 (fallback: ~9 px/caractere)."""
    texto = str(texto)
    if _FONTE_MEDIDA is None:
        return len(texto) * 9
    return _FONTE_MEDIDA.getlength(texto)


def _largura_coluna(textos):
    """Converte a maior largura (px) em unidade de coluna do Excel (autofit)."""
    maxpx = max((_px(t) for t in textos), default=0)
    return (maxpx + 9) / 7.0          # +9px de folga; 7px = largura do dígito padrão do Excel


def _ajustar_larguras(ws, col, linhas, cap=55):
    """Largura por coluna ajustada ao conteúdo REAL (mimetiza o autofit do duplo-clique,
    medindo a fonte Arial 12). 'code' fica menor; 'imagem' estreita; a oculta fica oculta."""
    for n, c in col.items():
        letra = get_column_letter(c)
        if n in COL_OCULTAS:
            ws.column_dimensions[letra].hidden = True
            continue
        if n == "imagem":
            ws.column_dimensions[letra].width = 5          # foto aparece no hover
            continue
        textos = [n] + [_texto_exibido(n, l.get(n)) for l in linhas]
        largura = min(_largura_coluna(textos), cap)
        if n == "code":
            largura = min(largura, 9)                      # menor que as demais (é só o ID)
        ws.column_dimensions[letra].width = round(largura, 1)


def _injetar_hover(caminho, imagens):
    """Pós-processa o xlsx: troca o fundo dos comentários (que o openpyxl já criou) pela
    IMAGEM da peça → a foto aparece ao passar o mouse. (Técnica VML legada.)

    imagens: lista de (row1based, col1based, caminho_jpg).
    """
    tmp = tempfile.mkdtemp()
    try:
        with zipfile.ZipFile(caminho) as z:
            z.extractall(tmp)
        # openpyxl gera o VML dos comentários como commentsDrawing*.vml (ou vmlDrawing*).
        vmls = glob.glob(os.path.join(tmp, "xl", "drawings", "*.vml"))
        if not vmls:
            raise RuntimeError("openpyxl não gerou o VML dos comentários")
        vml_path = vmls[0]
        with open(vml_path, encoding="utf-8") as f:
            vml = f.read()

        # os prefixos de namespace variam por versão (o/v/x ou ns0/ns1/ns2) → detecta
        def pfx(uri, padrao):
            m = re.search(rf'xmlns:(\w+)="{re.escape(uri)}"', vml)
            return m.group(1) if m else padrao
        VP = pfx("urn:schemas-microsoft-com:vml", "v")             # shape, fill
        OP = pfx("urn:schemas-microsoft-com:office:office", "o")   # relid
        XP = pfx("urn:schemas-microsoft-com:office:excel", "x")    # Row, Column

        media = os.path.join(tmp, "xl", "media")
        os.makedirs(media, exist_ok=True)
        rels, mapa = [], {}
        for n, (row, coln, img) in enumerate(imagens, start=1):
            relid, mf = f"rIdHov{n}", f"hover{n}.jpeg"
            shutil.copy(img, os.path.join(media, mf))
            rels.append((relid, mf))
            mapa[(row - 1, coln - 1)] = relid     # ClientData Row/Column são 0-based

        # troca o fundo de cada shape pela imagem (casado por linha/coluna)
        partes, saida = vml.split(f"</{VP}:shape>"), []
        for parte in partes:
            if f"<{VP}:shape " in parte or f"<{VP}:shape>" in parte:
                mr = re.search(rf"<{XP}:Row>(\d+)</{XP}:Row>", parte)
                mc = re.search(rf"<{XP}:Column>(\d+)</{XP}:Column>", parte)
                if mr and mc and (int(mr.group(1)), int(mc.group(1))) in mapa:
                    relid = mapa[(int(mr.group(1)), int(mc.group(1)))]
                    fill = (f'<{VP}:fill {OP}:relid="{relid}" {OP}:title="" recolor="t" '
                            f'rotate="t" type="frame"/>')
                    if re.search(rf"<{VP}:fill[^>]*/>", parte):
                        parte = re.sub(rf"<{VP}:fill[^>]*/>", fill, parte, count=1)
                    elif re.search(rf"<{VP}:fill[^>]*>.*?</{VP}:fill>", parte, flags=re.S):
                        parte = re.sub(rf"<{VP}:fill[^>]*>.*?</{VP}:fill>", fill, parte,
                                       count=1, flags=re.S)
                    else:
                        parte = parte.replace(f"<{VP}:shadow", fill + f"<{VP}:shadow", 1)
                saida.append(parte + f"</{VP}:shape>")
            else:
                saida.append(parte)
        with open(vml_path, "w", encoding="utf-8") as f:
            f.write("".join(saida))

        # rels do vml → imagens
        reldir = os.path.join(os.path.dirname(vml_path), "_rels")
        os.makedirs(reldir, exist_ok=True)
        relxml = ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                  '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
                  + "".join(f'<Relationship Id="{rid}" Type="http://schemas.openxmlformats.org/'
                            f'officeDocument/2006/relationships/image" Target="../media/{mf}"/>'
                            for rid, mf in rels)
                  + '</Relationships>')
        with open(os.path.join(reldir, os.path.basename(vml_path) + ".rels"), "w",
                  encoding="utf-8") as f:
            f.write(relxml)

        # content types: garante o tipo jpeg
        ctp = os.path.join(tmp, "[Content_Types].xml")
        with open(ctp, encoding="utf-8") as f:
            ct = f.read()
        if 'Extension="jpeg"' not in ct and 'Extension="jpg"' not in ct:
            ct = ct.replace("</Types>", '<Default Extension="jpeg" ContentType="image/jpeg"/></Types>')
            with open(ctp, "w", encoding="utf-8") as f:
                f.write(ct)

        # reempacota
        out = caminho + ".tmp"
        with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as z:
            for root, _, files in os.walk(tmp):
                for fn in files:
                    full = os.path.join(root, fn)
                    z.write(full, os.path.relpath(full, tmp).replace(os.sep, "/"))
        shutil.move(out, caminho)
    finally:
        shutil.rmtree(tmp, ignore_errors=True)
